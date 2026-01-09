# app/routes/inventory.py

import json
from typing import Any, Dict, List, Optional
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from app.db import get_session
from app.models.inventory import InventoryItem, InventoryRead
from app.models.product import Product
from app.models.salon import Salon

router = APIRouter(prefix="/inventory", tags=["inventory"])


@router.get(
    "",
    response_model=List[InventoryRead],
    summary="Lister l’inventaire par salon & produit",
)
def list_inventory(
    salon_id: Optional[int] = None,
    product_id: Optional[int] = None,
    session: Session = Depends(get_session),
) -> List[InventoryRead]:
    """
    Retourne l’inventaire.

    - Sans paramètre : tous les enregistrements
    - Avec salon_id : filtre par salon
    - Avec product_id : filtre par produit
    """
    stmt = select(InventoryItem)
    if salon_id is not None:
        stmt = stmt.where(InventoryItem.salon_id == salon_id)
    if product_id is not None:
        stmt = stmt.where(InventoryItem.product_id == product_id)

    return session.exec(stmt).all()


@router.get(
    "/view",
    summary="Vue lisible de l’inventaire (JOIN Product)",
)
def inventory_view(
    salon_id: Optional[int] = None,
    product_id: Optional[int] = None,
    session: Session = Depends(get_session),
) -> List[Dict[str, Any]]:
    """
    Vue lisible de l’inventaire (JOIN InventoryItem -> Product)
    Retourne: sku + name + quantity + options + price (si dispo).
    """
    stmt = select(InventoryItem, Product).join(Product, Product.id == InventoryItem.product_id)

    if salon_id is not None:
        stmt = stmt.where(InventoryItem.salon_id == salon_id)
    if product_id is not None:
        stmt = stmt.where(InventoryItem.product_id == product_id)

    rows = session.exec(stmt).all()

    out: List[Dict[str, Any]] = []
    for inv, prod in rows:
        out.append(
            {
                "inventory_id": inv.id,
                "salon_id": inv.salon_id,
                "product_id": inv.product_id,
                "quantity": inv.quantity,
                "sku": prod.sku,
                "name": prod.name,
                "price": getattr(prod, "price", None),
                "options": getattr(prod, "options", None),
                "wix_id": getattr(prod, "wix_id", None),
                "is_active": getattr(prod, "active", None),
            }
        )

    return out


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


@router.get(
    "/export.xlsx",
    summary="Exporter l’inventaire Excel (salons + catégories + filtres)",
)
def export_inventory_xlsx(
    salon_id: Optional[int] = None,
    include_zero: bool = False,
    category: Optional[str] = None,   # ✅ filtre catégorie unique
    session: Session = Depends(get_session),
):
    """
    Excel:
    - Sans category: 1 feuille par salon + 1 feuille par catégorie + Summary
    - Avec category: export filtré sur cette catégorie (dans options.categories)
    """
    # salons à exporter
    salons_stmt = select(Salon).where(Salon.is_active == True)  # noqa: E712
    if salon_id is not None:
        salons_stmt = select(Salon).where(Salon.id == salon_id)
    salons = session.exec(salons_stmt).all()

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "salon_id",
        "salon",
        "sku",
        "name",
        "quantity",
        "price",
        "value",
        "wix_id",
        "wix_variant_id",
        "categories",
        "choices",
        "vendor_sku",
    ]

    def extract_opts(prod: Product):
        opts = getattr(prod, "options", None) or {}
        wix_variant_id_val = None
        choices = None
        vendor_sku = None
        cats = []
        if isinstance(opts, dict):
            wix_variant_id_val = opts.get("wix_variant_id")
            choices = opts.get("choices")
            vendor_sku = opts.get("vendor_sku")
            cats = opts.get("categories") or []
        if not isinstance(cats, list):
            cats = []
        return wix_variant_id_val, choices, vendor_sku, cats

    def match_category(cats: List[str]) -> bool:
        if not category:
            return True
        want = category.strip().lower()
        return any(str(c).strip().lower() == want for c in cats)

    # Summary par SKU (global)
    summary: Dict[str, Dict[str, Any]] = {}

    # Accumulateur par catégorie (pour créer des feuilles "CAT - ...")
    by_category: Dict[str, List[List[Any]]] = {}

    # --- Feuilles par salon ---
    for s in salons:
        title = (s.name or f"Salon {s.id}")[:31]
        ws = wb.create_sheet(title=title)
        ws.append(headers)

        stmt = (
            select(InventoryItem, Product)
            .join(Product, Product.id == InventoryItem.product_id)
            .where(InventoryItem.salon_id == s.id)
        )
        rows = session.exec(stmt).all()

        for inv, prod in rows:
            qty = int(inv.quantity or 0)
            if (not include_zero) and qty == 0:
                continue

            price = float(getattr(prod, "price", 0) or 0)
            value = qty * price

            wix_variant_id_val, choices, vendor_sku, cats = extract_opts(prod)

            # filtre catégorie
            if not match_category(cats):
                continue

            # stringify pour Excel
            try:
                choices_str = json.dumps(choices, ensure_ascii=False) if choices is not None else ""
            except Exception:
                choices_str = str(choices) if choices is not None else ""

            cats_str = "; ".join([str(c) for c in cats]) if cats else ""
            vendor_sku_str = "" if vendor_sku is None else str(vendor_sku)
            wix_variant_id_str = "" if wix_variant_id_val is None else str(wix_variant_id_val)
            wix_id_str = "" if getattr(prod, "wix_id", None) is None else str(getattr(prod, "wix_id", None))

            row_out = [
                s.id,
                s.name,
                prod.sku,
                prod.name,
                qty,
                price,
                value,
                wix_id_str,
                wix_variant_id_str,
                cats_str,
                choices_str,
                vendor_sku_str,
            ]

            ws.append(row_out)

            # --- Summary SKU ---
            sku_key = prod.sku or ""
            if sku_key:
                agg = summary.get(sku_key) or {"sku": sku_key, "name": prod.name, "qty": 0, "price": price, "value": 0.0}
                agg["qty"] += qty
                agg["value"] += value
                if prod.name:
                    agg["name"] = prod.name
                agg["price"] = price
                summary[sku_key] = agg

            # --- Feuilles catégories (uniquement si pas de filtre category) ---
            if not category:
                for c in cats or ["(Sans catégorie)"]:
                    cname = str(c).strip() or "(Sans catégorie)"
                    by_category.setdefault(cname, []).append(row_out)

        _autosize(ws)

    # --- Feuilles par catégorie (quand on exporte tout) ---
    if not category:
        for cname in sorted(by_category.keys()):
            sheet_name = f"CAT - {cname}"[:31]
            ws_cat = wb.create_sheet(title=sheet_name)
            ws_cat.append(headers)
            for r in by_category[cname]:
                ws_cat.append(r)
            _autosize(ws_cat)

    # --- Summary ---
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.append(["sku", "name", "total_qty", "price", "total_value"])
    for sku_key in sorted(summary.keys()):
        a = summary[sku_key]
        ws_sum.append([a["sku"], a["name"], a["qty"], a["price"], a["value"]])
    _autosize(ws_sum)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    if category:
        filename = f"luxura_inventory_{category}_{stamp}.xlsx"
    else:
        filename = f"luxura_inventory_{stamp}.xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
